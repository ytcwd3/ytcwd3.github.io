#!/usr/bin/env python3
"""
Excel数据导出工具 - 完全匹配原始数据
根据每个Sheet的特性智能处理
"""
import openpyxl
import json

def clean_value(val):
    """清洗单元格值，保持原始格式"""
    if val is None:
        return ''
    val = str(val).strip()
    if val == 'None' or val == 'none':
        return ''
    return val

def is_header_row(row):
    """判断是否是标题行"""
    if not row[0]:
        return False
    first_val = str(row[0]).strip()
    # 标题行特征
    if '游戏名' in first_val or '母分类' in first_val:
        return True
    return False

def export_excel_to_json(excel_path, output_path):
    """导出Excel到JSON，保持原始数据结构"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_games = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n处理 Sheet: {sheet_name} (共 {ws.max_row} 行)")

        # 检测表头行
        header_row_idx = None
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            if is_header_row(row):
                header_row_idx = row_idx
                print(f"  发现标题行: 第{row_idx}行")
                break

        # 确定数据起始行
        data_start = header_row_idx + 1 if header_row_idx else 1

        # 根据Sheet名称设置默认分类
        default_category = []
        default_subcategory = []

        if 'NS' in sheet_name:
            default_category = ['任天堂']
            default_subcategory = ['NS']
            print(f"  默认分类: 任天堂, 子分类: NS")
        elif 'other' in sheet_name.lower():
            default_category = ['其他平台']
            print(f"  默认分类: 其他平台")

        # 确定列映射
        if sheet_name == 'other':
            col_map = {
                'name': 0, 'platform1': 1, 'platform2': 2,
                'subcat1': 3, 'subcat2': 4, 'code': 5,
                'quark': 6, 'baidu': 7, 'thunder': 8,
                'unzip': 9, 'date': 10
            }
        else:
            col_map = {
                'name': 0, 'platform1': 1, 'platform2': 2,
                'subcat1': 3, 'subcat2': 4, 'code': 5,
                'quark': 6, 'baidu': 7, 'thunder': 8,
                'unzip': 9, 'date': 10
            }

        # 导出数据
        count = 0
        for row_idx in range(data_start, ws.max_row + 1):
            row = [cell.value for cell in ws[row_idx]]

            name = clean_value(row[col_map['name']])

            # 跳过空行或无效行
            if not name or name == '':
                continue

            # 构建category数组
            platforms = []
            p1 = clean_value(row[col_map['platform1']])
            p2 = clean_value(row[col_map['platform2']])

            if p1:
                platforms.append(p1)
            elif default_category:
                # 如果平台1为空，使用默认值
                platforms.extend(default_category)

            if p2:
                platforms.append(p2)

            # 构建subcategory数组
            subcats = []
            s1 = clean_value(row[col_map['subcat1']])
            s2 = clean_value(row[col_map['subcat2']])

            if s1:
                subcats.append(s1)
            elif default_subcategory:
                # 如果子分类1为空，使用默认值
                subcats.extend(default_subcategory)

            if s2:
                subcats.append(s2)

            # 提取码处理
            code = row[col_map['code']]
            if isinstance(code, (int, float)):
                code = str(int(code))
            else:
                code = clean_value(code)

            # 解压码处理
            unzipcode = clean_value(row[col_map['unzip']])

            game = {
                'name': name,
                'category': platforms if platforms else [],
                'subCategory': subcats if subcats else [],
                'code': code,
                'unzipCode': unzipcode,
                'quarkPan': clean_value(row[col_map['quark']]),
                'baiduPan': clean_value(row[col_map['baidu']]),
                'thunderPan': clean_value(row[col_map['thunder']]),
                'updateDate': clean_value(row[col_map['date']])
            }

            all_games.append(game)
            count += 1

        print(f"  导出: {count} 条")

    # 保存JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_games, f, ensure_ascii=False, indent=2)

    print(f"\n总计导出: {len(all_games)} 条数据")
    print(f"保存至: {output_path}")

    return all_games

def verify_data(data, excel_path):
    """验证数据完整性"""
    print("\n=== 数据验证 ===")

    # 检查空分类
    empty_cat = [g for g in data if not g['category']]
    if empty_cat:
        print(f"警告: {len(empty_cat)} 条数据category为空")
        for g in empty_cat[:3]:
            print(f"  - {g['name']}")

    # 统计各平台
    from collections import Counter
    all_cats = []
    all_subcats = []
    for g in data:
        all_cats.extend(g['category'])
        all_subcats.extend(g['subCategory'])

    print("\n各平台数量:")
    for cat, count in Counter(all_cats).most_common():
        print(f"  {cat}: {count}")

    print(f"\n总数据量: {len(data)}")

if __name__ == '__main__':
    excel_path = '/Users/whitefox/Desktop/work/ytcwd3.github.io/自己分.xlsx'
    output_path = '/Users/whitefox/Desktop/work/ytcwd3.github.io/data/supabase_import.json'

    games = export_excel_to_json(excel_path, output_path)
    verify_data(games, excel_path)

    # 显示示例
    print("\n=== 示例数据 ===")
    for g in games[:2]:
        print(json.dumps(g, ensure_ascii=False, indent=2))
